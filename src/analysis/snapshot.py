import re
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table
from pymongo import DESCENDING

from log import LogHandler
from src.analysis import conn
from src.analysis.setting import REPORT_DIR, TABLE_STYLE, COLOR_RULE, PERCENT_FORMAT, COMMA0_FORMAT, DATE_FORMAT, \
    PERCENT0_FORMAT
from src.api import get_peak_start_date, get_price, get_roll_yield
from src.data.future.setting import CODE2NAME_MAP

# from src.util import count_percentile

log = LogHandler('analysis.log')


def get_instrument_symbols(by='amount', threshold=1e9, instrument='future'):
    """
    根据流动性选择交易品种，30天内的平均值作为选择标准，由于节假日原因，取平均的数据个数可能不一样。
    :param instrument:
    :param by:过滤的因子
    :param threshold:过滤的阈值
    :return:
    """
    # conn = connect_mongo(db='quote', username=DATA_ANALYST, password=ANALYST_PWD)

    pipeline = [
        {
            '$match': {
                'datetime': {
                    '$gt': datetime.today() - timedelta(30)  # 选取30天内有交易的品种
                }
            }
        }, {
            '$sort': {
                'datetime': DESCENDING
            }
        }, {
            '$group': {
                '_id': '$symbol',
                'datetime': {
                    '$first': '$datetime'
                },
                by: {
                    '$avg': '$' + by
                }
            }
        }, {
            '$match': {
                by: {
                    '$gt': threshold
                }
            }
        }, {
            '$sort': {
                by: DESCENDING
            }
        }
    ]

    match_stage = pipeline[0]['$match']
    cursor = conn['index']
    if instrument == 'future':
        match_stage['symbol'] = {'$regex': re.compile(r"88$")}
    else:
        pass

    symbol_list = list(cursor.aggregate(pipeline))
    # 剔除某一天超过阈值的
    if len(symbol_list) == 0:
        log.info('None of snapshot dates return!')
        return symbol_list
    return [symbol['_id'] for symbol in symbol_list]
    # return symbol_list


def get_last_price(symbol=None, instrument='index', frequency='d', fields='close'):
    """
        获取最新的行情数据
    :param symbol: 合约代码，symbol, symbol list, 只支持同种类。获取tick数据时，只支持单个symbol
    :param instrument:   行情数据类型 ['future', 'option', 'stock', 'bond', 'convertible'] 以及 index
    :param frequency:   历史数据的频率, 默认为'd', 只支持日线级别以上数据。'5m'代表5分钟线。可支持期货tick级别数据获取，此时频率为'tick'
    :param fields:      字段名称
    :return:
        传入一个symbol，多个fields，函数会返回一个pandas DataFrame
        传入一个symbol，一个field，函数会返回pandas Series
        传入多个symbol，一个field，函数会返回一个pandas DataFrame
        传入多个symbol，函数会返回一个multiIndex DataFrame
    """
    # 连接数据库
    # conn = connect_mongo(db='quote', username=DATA_ANALYST, password=ANALYST_PWD)

    cursor = conn[instrument]

    pipeline = [
        {
            '$match': {
                'symbol': {
                    '$regex': re.compile(r"88$")
                }
            }
        }, {
            '$sort': {
                'datetime': DESCENDING
            }
        }, {
            '$group': {
                '_id': '$symbol',
                'datetime': {
                    '$first': '$datetime'
                },
                'close': {
                    '$first': '$close'
                },
                'amount': {
                    '$first': '$amount'
                }
            }
        }, {
            '$sort': {
                'amount': DESCENDING
            }
        }
    ]

    match_stage = pipeline[0]['$match']
    if isinstance(symbol, list):
        match_stage['symbol'] = {'$in': symbol}
    elif isinstance(symbol, str):
        match_stage['symbol'] = symbol
    else:
        log.debug('Search all instruments snapshot start datetime!')

    last_prices = cursor.aggregate(pipeline)

    last_prices_df = pd.DataFrame(list(last_prices))
    last_prices_df.rename(columns={'_id': 'symbol'}, inplace=True)
    return last_prices_df


def get_future_snapshot(symbol=None, instrument='index', threshold=1e9):
    if symbol is None:
        symbol = get_instrument_symbols(by='amount', threshold=threshold)

    start_df = get_peak_start_date(symbol=symbol)

    start_df['start_date'] = start_df['start_date'].fillna(datetime(1970, 1, 1))

    # price status
    # columns = ['name', 'percentile', 'wave_rt',
    #            'amount', 'close', 'highest', 'lowest',
    #            'contract', 'start_date', 'datetime']

    snapshot_df = pd.DataFrame()

    index_cursor = conn[instrument]

    for row in start_df.itertuples():
        symbol, start_date = row.symbol, row.start_date
        filter_dict = {
            'symbol': symbol,
            'datetime': {
                '$gte': start_date
            }
        }

        hq = index_cursor.find(filter_dict).sort([('datetime', DESCENDING)])

        hq_df = pd.DataFrame(list(hq))

        hq_df = hq_df[hq_df['low'] > 0]  # 历史成交量可能为0，没有成交价格

        last_hq = hq_df.iloc[0]

        code = last_hq['code']
        snapshot_df.loc[code, 'name'] = CODE2NAME_MAP.get(code, code)

        close = last_hq['close']
        highest = hq_df['high'].max()
        lowest = hq_df['low'].min()
        snapshot_df.loc[code, 'wave_rt'] = (close - lowest) / (highest - lowest)

        snapshot_df.loc[code, 'amount'] = last_hq['amount'] / 1e8
        snapshot_df.loc[code, 'close'] = close
        snapshot_df.loc[code, 'highest'] = highest
        snapshot_df.loc[code, 'lowest'] = lowest
        snapshot_df.loc[code, 'contract'] = last_hq['contract']
        snapshot_df.loc[code, 'start_date'] = start_date
        start_date = snapshot_df.loc[code, 'datetime'] = last_hq['datetime']

        filter_dict = {
            'symbol': {"$regex": "^" + code + "[7-9]{2}$"},
            'datetime': {
                '$gte': start_date
            }
        }

        projection = {
            "_id": 0,
            "symbol": 1,
            "datetime": 1,
            "close": 1
        }

        # 只需要取最后一天的数据
        hq = index_cursor.find(filter_dict, projection=projection)

        hq_df = pd.DataFrame(list(hq))
        hq_df = hq_df.pivot(index='datetime', columns='symbol', values='close')

        if len(hq_df.columns) != 3:
            log.warning('{} hq data are not enough!'.format(code))

        spot_cursor = conn['spot_price']
        filter_dict = {"code": code,
                       'datetime': {
                           '$gte': start_date
                       }}
        projection = {"_id": 0, "datetime": 1, "spot": 1}
        spot = spot_cursor.find(filter_dict, projection=projection)
        spot_df = pd.DataFrame(list(spot))
        if spot_df.empty:
            yield_df = hq_df
            yield_df.columns = ['deliver', 'domain', 'far_month']
            yield_df['deliver_basis'] = np.nan
        else:
            spot_df.set_index('datetime', inplace=True)
            yield_df = pd.concat([spot_df, hq_df], axis=1)
            yield_df = yield_df.dropna()
            yield_df.columns = ['spot', 'deliver', 'domain', 'far_month']
            yield_df['deliver_basis'] = (yield_df['deliver'] / yield_df['spot'] - 1)
            yield_df['domain_basis'] = (yield_df['domain'] / yield_df['spot'] - 1)
            yield_df['far_month_basis'] = (yield_df['far_month'] / yield_df['spot'] - 1)

            last_yield = yield_df.iloc[-1]

            snapshot_df.loc[code, 'domain_basis'] = last_yield['domain_basis']
            snapshot_df.loc[code, 'far_month_basis'] = last_yield['far_month_basis']

        yield_df['nearby_yield'] = (yield_df['domain'] / yield_df['deliver'] - 1)
        yield_df['far_month_yield'] = (yield_df['far_month'] / yield_df['domain'] - 1)

        last_yield = yield_df.iloc[-1]

        snapshot_df.loc[code, 'deliver_basis'] = last_yield['deliver_basis']

        snapshot_df.loc[code, 'nearby_yield'] = last_yield['nearby_yield']
        snapshot_df.loc[code, 'far_month_yield'] = last_yield['far_month_yield']

    snapshot_df.sort_values(by='wave_rt', inplace=True)
    return snapshot_df


def generate_future_report():
    report_dir = REPORT_DIR / 'future'
    future_filepath = report_dir / 'snapshot{}.xlsx'.format(datetime.today().strftime('%Y%m%d'))

    report_dir.mkdir(parents=True, exist_ok=True)

    if future_filepath.exists():
        # snapshot_filepath.unlink()
        # wb = load_workbook(future_filepath)
        print("{} is exists!".format(future_filepath.name))

    wb = Workbook()

    ws = wb.active
    ws.title = 'snapshot'
    snapshot_df = get_future_snapshot(threshold=1e11)

    df = snapshot_df[
        ['name', 'wave_rt', 'deliver_basis', 'domain_basis', 'far_month_basis', 'nearby_yield', 'far_month_yield',
         'amount', 'highest', 'lowest', 'close', 'contract', 'start_date', 'datetime']]

    columns = ['交易品种', '价格比例', '交割基差', '主力基差', '远月基差', '交割换月价差', '远月换月价差',
               '成交金额', '最高价', '最低价', '收盘价', '主力合约', '开始日期', '更新日期']
    ws.append(columns)
    values = df.values.tolist()
    for value in values:
        ws.append(value)

    max_row = len(snapshot_df) + 1
    bottom_right = ws.cell(row=max_row,
                           column=len(snapshot_df.columns)).coordinate
    snapshot_range = "A1:{}".format(bottom_right)

    first_row = ws.row_dimensions[0]
    first_row.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.iter_cols(min_col=1, max_col=len(snapshot_df.columns), min_row=1, max_row=len(snapshot_df) + 1):
        for cell in col:
            i = cell.column
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if i in range(1, 8):
                cell.number_format = PERCENT_FORMAT
            elif i in range(8, 12):
                cell.number_format = COMMA0_FORMAT
            elif i in [13, 14]:
                cell.number_format = DATE_FORMAT

    tab = Table(displayName='snapshot', ref=snapshot_range)

    tab.tableStyleInfo = TABLE_STYLE

    ws.add_table(tab)

    columns_formatting = ['B', 'D', 'E', 'G']

    for col in columns_formatting:
        range_formatting = '{0}2:{0}{1}'.format(col, max_row)
        ws.conditional_formatting.add(range_formatting, COLOR_RULE)

    wb.save(future_filepath)


def generate_future_detail_report(code, start_date):
    """
    生成单个交易品种的分析
    :param code: 期货简码
    :param start_date: 分析数据的起始日期
    :return:
    """

    wb = Workbook()
    ws = wb.active
    ws.title = code

    symbol = code + '88'
    # hq_df = get_price(symbol=symbol, start_date=start_date)

    # ws_hq = wb.create_sheet(title='hq')
    #
    # for r in dataframe_to_rows(hq_df, index=False, header=True):
    #     ws_hq.append(r)

    basis_df = get_roll_yield(code=code, start_date=start_date)
    basis_df.reset_index(inplace=True)

    last_domain_close = basis_df[symbol].iloc[-1]
    last_domain_basis = basis_df.domain_basis.iloc[-1]
    basis_df['last_domain_close'] = last_domain_close
    basis_df['last_domain_basis'] = last_domain_basis

    ws_basis = wb.create_sheet(title='basis')

    for r in dataframe_to_rows(basis_df, index=False, header=True):
        ws_basis.append(r)

    # 基差图
    chart_basis = LineChart()
    chart_basis.title = "Basis"
    chart_basis.height = 15  # default is 7.5
    chart_basis.width = 45  # default is 15
    chart_basis.style = 2
    chart_basis.y_axis.title = "%"
    chart_basis.y_axis.crossAx = 500
    chart_basis.y_axis.number_format = PERCENT0_FORMAT
    chart_basis.x_axis = DateAxis(crossAx=100)
    chart_basis.x_axis.number_format = DATE_FORMAT
    chart_basis.x_axis.majorTimeUnit = "days"
    chart_basis.x_axis.title = "Date"

    data = Reference(ws_basis, min_col=6, min_row=1, max_col=8, max_row=ws_basis.max_row)
    chart_basis.add_data(data, titles_from_data=True)
    values = Reference(ws_basis, min_col=12, min_row=1, max_row=ws_basis.max_row)
    series = Series(values, title_from_data=True)
    chart_basis.series.append(series)
    dates = Reference(ws_basis, min_col=1, min_row=2, max_row=ws_basis.max_row)
    chart_basis.set_categories(dates)

    # 收盘价图
    chart_close = LineChart()
    chart_close.y_axis.title = "close"
    chart_close.y_axis.axId = 200
    chart_close.y_axis.majorGridlines = None

    close_min_str = str(int(basis_df[symbol].min()))
    y_axis_min = int(close_min_str[0] + '0' * (len(close_min_str) - 1))
    chart_close.y_axis.scaling.min = y_axis_min

    values = Reference(ws_basis, min_col=4, min_row=1, max_row=ws_basis.max_row)
    chart_close.add_data(values, titles_from_data=True)
    s = chart_close.series[0]
    s.graphicalProperties.line.dashStyle = "sysDash"
    s.graphicalProperties.line.width = 3.0
    chart_basis.y_axis.crosses = 'max'
    chart_basis += chart_close
    ws.add_chart(chart_basis, "B2")

    # ws.add_chart(chart_close, "B32")

    report_dir = REPORT_DIR / 'future'
    future_filepath = report_dir / "{}{}.xlsx".format(code, datetime.today().strftime('%Y%m%d'))
    wb.save(future_filepath)


if __name__ == '__main__':
    # symbols = ['A88', 'Y88', 'ME88']
    # symbols = get_instrument_symbols(by='amount')
    # start_dates = get_peak_start_date(symbol=symbols)
    # print(start_dates)
    # last_price_df = get_last_price(symbol=symbols)
    # snapshot_df = get_future_snapshot(threshold=1e11)
    # with pd.ExcelWriter('output.xlsx') as writer:
    #     snapshot_df.to_excel(writer, sheet_name='snapshot')
    # block_status_df.to_excel(writer, sheet_name='block_status')
    # roll_yield_df.to_excel(writer, sheet_name='roll_yield')
    # print(last_price_df)
    code = 'I'
    start_date = datetime(2016, 11, 24)
    generate_future_detail_report(code=code, start_date=start_date)
    # generate_future_report()
