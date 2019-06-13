import numpy as np
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart import Series

from src.analysis.setting import COMMA0_FORMAT, PERCENT_FORMAT


def histogram(ws, series: np.ndarray, bins=50):
    """

    :param ws: 储存数据的worksheet
    :param series: 画直方图的数据
    :param bins: 分段的个数
    :return: chart
    """
    hist, bin_edges = np.histogram(series, bins)
    count = np.insert(hist, 0, hist[0])
    max_row = bins + 1
    current = np.repeat(series[-1], max_row)
    max_number = np.linspace(hist.min(), hist.max(), max_row, endpoint=True)
    data = np.vstack((bin_edges, count, current, max_number)).transpose()

    for r in data:
        ws.append(r.tolist())

    chart = ScatterChart()
    chart.width = 22  # default is 15
    chart.height = 15  # default is 7.5
    chart.style = 2
    chart.title = "Basis Distribution"
    chart.y_axis.title = 'Count'
    chart.x_axis.majorGridlines = None
    chart.y_axis.number_format = COMMA0_FORMAT
    chart.x_axis.title = 'Bin'
    chart.x_axis.number_format = PERCENT_FORMAT

    yvalues = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    xvalues = Reference(ws, min_col=1, min_row=1, max_row=ws.max_row)
    series = Series(values=yvalues, xvalues=xvalues, title='Count')
    series.smooth = True
    chart.series.append(series)
    # chart.series[-1].smooth = True

    yvalues = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    xvalues = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
    series = Series(values=yvalues, xvalues=xvalues, title='Current')
    chart.series.append(series)

    return chart
